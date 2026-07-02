"""
build_excel.py
==============
Builds the professional Excel summary workbook for the Inventory
Optimisation project. Reads CSVs from outputs/ directory.

Sheets:
  1. Executive Summary
  2. ABC Classification
  3. XYZ Classification
  4. ABC-XYZ Matrix & Strategy
  5. Reorder Points (Top 80)
  6. Rationalisation List
  7. Methodology & Limitations
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

OUT = Path("outputs")

# ── Colour palette (ARGB 8-char) ─────────────────────────────────────────────
NAVY    = "FF1A3C5E"
CORAL   = "FFC0392B"
AMBER   = "FFE67E22"
GREEN   = "FF27AE60"
SKY     = "FF2980B9"
SILVER  = "FFBDC3C7"
DGREY   = "FF2C3E50"
WHITE   = "FFFFFFFF"
LTGREY  = "FFECF0F1"
MIDGREY = "FF95A5A6"
OFFWHITE= "FFFAFAFA"
PURPLE  = "FF8E44AD"
RED_BG  = "FFFDEDEC"
AMBER_BG= "FFFEF9E7"
GREEN_BG= "FFEAFAF1"
BLUE_IN = "FF2471A3"   # blue = hardcoded input cells

# ── Helper functions ──────────────────────────────────────────────────────────
def fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def font(size=11, color="FF000000", bold=False, italic=False):
    return Font(name="Arial", size=size, color=color, bold=bold, italic=italic)

_thin = Side(style="thin", color="FFBDC3C7")
_border = Border(top=_thin, bottom=_thin, left=_thin, right=_thin)

def ba(cell):
    """Apply border-all to cell."""
    cell.border = _border

def w(ws, row, col, value, sz=11, color="FF000000", bold=False, italic=False,
      bg=None, align="center", wrap=False, nf=None):
    """Write a formatted cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(sz, color, bold, italic)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    ba(cell)
    if nf:
        cell.number_format = nf
    return cell

def banner(ws, title: str, subtitle: str, cols: int = 9):
    """Full-width title + subtitle banner."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = font(15, WHITE, bold=True)
    t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = font(9, DGREY, italic=True)
    s.fill = fill(LTGREY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 16

def section(ws, row: int, text: str, cols: int = 9, bg=NAVY, fg=WHITE):
    """Full-width section header."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(10, fg, bold=True)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

def header_row(ws, row: int, headers: list, widths: list, bg=NAVY):
    """Column header row."""
    for i, (h, wid) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font(10, WHITE, bold=True)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ba(c)
        ws.column_dimensions[get_column_letter(i)].width = wid
    ws.row_dimensions[row].height = 28

def spacer(ws, row: int, cols: int = 9):
    """Empty spacer row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 7

def note_row(ws, row: int, text: str, cols: int = 9, bg=AMBER_BG, color=DGREY):
    """Full-width note / callout row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(10, color)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1, wrap_text=True)
    ba(c)
    ws.row_dimensions[row].height = 30

# ── Load outputs ──────────────────────────────────────────────────────────────
def load(filename: str) -> pd.DataFrame:
    path = OUT / filename
    if not path.exists():
        print(f"  WARNING: {filename} not found — run analysis.py first")
        return pd.DataFrame()
    return pd.read_csv(path)

sku_master  = load("sku_master.csv")
abc_xyz_mat = load("abc_xyz_matrix.csv")
rop_df      = load("reorder_points.csv")
ration_df   = load("rationalisation_list.csv")
summary_df  = load("summary_stats.csv")

if sku_master.empty:
    print("ERROR: sku_master.csv is empty. Ensure analysis.py ran successfully.")
    sys.exit(1)

# Pre-compute summaries
total_rev   = sku_master["total_revenue"].sum()
total_skus  = len(sku_master)

abc_sum = (
    sku_master.groupby("ABC")
    .agg(sku_count=("StockCode","count"), revenue=("total_revenue","sum"))
    .reindex(["A","B","C"]).reset_index()
)
abc_sum["pct_skus"] = abc_sum["sku_count"] / total_skus
abc_sum["pct_rev"]  = abc_sum["revenue"]   / total_rev

xyz_sum = (
    sku_master.groupby("XYZ")
    .agg(sku_count=("StockCode","count"), cv_mean=("cv","mean"))
    .reindex(["X","Y","Z"]).reset_index()
)
xyz_sum["pct_skus"] = xyz_sum["sku_count"] / total_skus

n_az_bz = len(sku_master[sku_master["ABC_XYZ"].isin(["AZ","BZ"])])
rev_az_bz = sku_master[sku_master["ABC_XYZ"].isin(["AZ","BZ"])]["total_revenue"].sum()
n_cz    = len(sku_master[sku_master["ABC_XYZ"] == "CZ"])

wb = Workbook()


# ════════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
banner(ws1,
    "Inventory Optimisation — UK Wholesale Distributor",
    "Dataset: UCI Online Retail (real UK wholesale, 525K+ transactions, Dec 2010–Dec 2011) · "
    "Analyst: Vaishnavi Bhor")

# KPI strip
kpis = [
    ("Total SKUs",       f"{total_skus:,}",          "After data cleaning",               SKY),
    ("Total Revenue",    f"£{total_rev/1e6:.1f}M",    "12-month period",                   NAVY),
    ("A-Class SKUs",     f"{int(abc_sum.loc[abc_sum['ABC']=='A','sku_count'].values[0]):,}", "Drive 80% of revenue", CORAL),
    ("AZ+BZ SKUs",       f"{n_az_bz:,}",              "High value, unpredictable demand",  CORAL),
    ("AZ+BZ Revenue",    f"£{rev_az_bz/1e6:.1f}M",    "At risk from stockouts",            AMBER),
    ("CZ SKUs",          f"{n_cz:,}",                 "Rationalisation candidates",        AMBER),
    ("Z-Class SKUs",     f"{int(xyz_sum.loc[xyz_sum['XYZ']=='Z','sku_count'].values[0]):,}", "Intermittent demand (CV>1)", AMBER),
    ("ROP Model",        "Top 80 SKUs",               "95-98% service level targets",      GREEN),
    ("Data Source",      "UCI + Real",                "525K+ real transactions",           SKY),
]
for i, (lbl, val, note_txt, col) in enumerate(kpis, 1):
    ws1.column_dimensions[get_column_letter(i)].width = 15
    for row_num, cell_val, is_main in [(4, lbl, False), (5, val, True), (6, note_txt, False)]:
        c = ws1.cell(row=row_num, column=i, value=cell_val)
        if row_num == 4:
            c.font = font(8, WHITE, bold=True)
            c.fill = fill(col)
        elif row_num == 5:
            c.font = font(14, col, bold=True)
            c.fill = fill(OFFWHITE)
        else:
            c.font = font(8, MIDGREY, italic=True)
            c.fill = fill(OFFWHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[4].height = 26
ws1.row_dimensions[5].height = 32
ws1.row_dimensions[6].height = 22

spacer(ws1, 7)
section(ws1, 8, "ANALYTICAL FRAMEWORK — McKINSEY ISSUE TREE (3 BRANCHES)")
branches = [
    ("Branch 1", "ABC Classification\n(Pareto by revenue)", "Which SKUs drive the most revenue? Apply differentiated service levels by class.", "Sheet: ABC Classification"),
    ("Branch 2", "XYZ Classification\n(Demand variability)", "How predictable is each SKU's demand? Uses Coefficient of Variation on weekly demand.", "Sheet: XYZ Classification"),
    ("Branch 3", "Combined ABC-XYZ\n(9-cell strategy matrix)", "Overlay classifications to assign strategic action per cell.", "Sheet: ABC-XYZ Matrix"),
    ("Branch 4", "Reorder Point Model\n(Top 80 SKUs)", "Calculate ROP with safety stock for highest-revenue SKUs. Service level by class.", "Sheet: Reorder Points"),
    ("Branch 5", "Rationalisation Analysis\n(CZ review)", "Identify CZ candidates for discontinuation. Apply loyal-account exception.", "Sheet: Rationalisation"),
]
header_row(ws1, 9, ["Branch", "Method", "Business Question", "Deliverable"], [12, 22, 44, 24])
for i, (branch, method, question, deliv) in enumerate(branches, 10):
    ws1.row_dimensions[i].height = 32
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws1, i, 1, branch,   10, CORAL,  bold=True, bg=bg)
    w(ws1, i, 2, method,   10, NAVY,   bold=True, bg=bg, align="left", wrap=True)
    w(ws1, i, 3, question, 10, bg=bg,  align="left", wrap=True)
    w(ws1, i, 4, deliv,     9, SKY,    italic=True, bg=bg, align="left")

spacer(ws1, 15)
section(ws1, 16, "KEY FINDINGS")
findings = [
    ("Finding 1", "Classic Pareto holds strongly",
     f"{int(abc_sum.loc[abc_sum['ABC']=='A','sku_count'].values[0]):,} A-class SKUs "
     f"({abc_sum.loc[abc_sum['ABC']=='A','pct_skus'].values[0]:.0%} of range) "
     f"drive 80% of £{total_rev/1e6:.1f}M annual revenue. "
     "Inventory investment should be heavily concentrated here."),
    ("Finding 2", "Demand is predominantly variable (Z-class dominant)",
     f"{int(xyz_sum.loc[xyz_sum['XYZ']=='Z','sku_count'].values[0]):,} SKUs "
     f"({xyz_sum.loc[xyz_sum['XYZ']=='Z','pct_skus'].values[0]:.0%}) have CV > 1.0. "
     "Pure statistical forecasting is insufficient for this assortment. "
     "Demand sensing and collaborative forecasting with key accounts is recommended."),
    ("Finding 3", f"AZ+BZ represents {n_az_bz:,} high-value, unpredictable SKUs",
     f"£{rev_az_bz:,.0f} in annual revenue sits in SKUs with unpredictable demand. "
     "These warrant management attention: safety stock alone is insufficient. "
     "Demand sensing, lead-time reduction, and closer customer collaboration are the levers."),
    ("Finding 4", "CZ rationalisation: most are protected by loyal accounts",
     f"Of {n_cz:,} CZ SKUs, only 19 are clear rationalisation candidates (£886 total revenue). "
     "The majority serve customers with repeat purchase patterns. "
     "The framework identifies the candidates — commercial judgment decides the action."),
]
header_row(ws1, 17, ["Finding", "Headline", "Evidence"], [12, 30, 58])
for i, (label, headline, evidence) in enumerate(findings, 18):
    ws1.row_dimensions[i].height = 38
    bg = RED_BG if i == 18 else AMBER_BG if i <= 19 else OFFWHITE
    w(ws1, i, 1, label,   10, CORAL, bold=True, bg=bg)
    w(ws1, i, 2, headline, 10, NAVY, bold=True, bg=bg, align="left", wrap=True)
    w(ws1, i, 3, evidence, 9, DGREY, bg=bg, align="left", wrap=True)


# ════════════════════════════════════════════════════════════════════
# SHEET 2 — ABC CLASSIFICATION
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("ABC Classification")
ws2.sheet_view.showGridLines = False
banner(ws2, "Branch 1: ABC Classification — Pareto Revenue Analysis",
       "A = top 80% revenue · B = 80-95% · C = 95-100% · Standard inventory Pareto (Vilfredo Pareto, 1896)")

section(ws2, 4, "ABC SUMMARY TABLE")
header_row(ws2, 5,
    ["Class", "SKU Count", "% of SKUs", "Annual Revenue (£)", "% of Revenue",
     "Avg Unit Price (£)", "Strategic Implication"],
    [8, 12, 12, 20, 14, 18, 44])

abc_colors = {"A": CORAL, "B": AMBER, "C": SKY}
implications = {
    "A": "Premium service (95-98% SL). Tight ROP. Weekly management review. No stockouts tolerated.",
    "B": "Standard service (90-95% SL). Monthly review. Moderate safety stock.",
    "C": "Lean. Minimal inventory. Quarterly review. Consider consolidation.",
}
for i, (_, row) in enumerate(abc_sum.iterrows(), 6):
    ws2.row_dimensions[i].height = 28
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    col = abc_colors.get(row["ABC"], DGREY)
    avg_price = sku_master[sku_master["ABC"] == row["ABC"]]["avg_unit_price"].mean()
    w(ws2, i, 1, row["ABC"],       12, col, bold=True, bg=bg)
    w(ws2, i, 2, row["sku_count"], 10, bg=bg, nf="#,##0")
    w(ws2, i, 3, row["pct_skus"],  10, bg=bg, nf="0.0%")
    w(ws2, i, 4, row["revenue"],   10, col, bold=True, bg=bg, nf="£#,##0")
    w(ws2, i, 5, row["pct_rev"],   10, col, bold=True, bg=bg, nf="0.0%")
    w(ws2, i, 6, avg_price,        10, bg=bg, nf="£#,##0.00")
    w(ws2, i, 7, implications.get(row["ABC"], ""), 9, DGREY, bg=bg, align="left", wrap=True)

spacer(ws2, 9)
section(ws2, 10, "TOP 30 A-CLASS SKUs (sorted by revenue)")
a_skus = sku_master[sku_master["ABC"] == "A"].nlargest(30, "total_revenue")
header_row(ws2, 11,
    ["Rank", "Stock Code", "Description", "Annual Revenue (£)", "Annual Units Sold",
     "Avg Price (£)", "XYZ Class", "ABC-XYZ"],
    [6, 12, 36, 18, 18, 14, 10, 10])
for i, (rank, (_, row)) in enumerate(enumerate(a_skus.iterrows(), 1), 12):
    ws2.row_dimensions[i].height = 22
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    xyz_col = {"X": GREEN, "Y": AMBER, "Z": CORAL}.get(row["XYZ"], DGREY)
    w(ws2, i, 1, rank,                   9, MIDGREY, bg=bg)
    w(ws2, i, 2, row["StockCode"],        9, NAVY, bold=True, bg=bg)
    w(ws2, i, 3, str(row["description"])[:50], 9, bg=bg, align="left")
    w(ws2, i, 4, row["total_revenue"],   10, CORAL, bold=True, bg=bg, nf="£#,##0.00")
    w(ws2, i, 5, row["total_qty_sold"],  10, bg=bg, nf="#,##0")
    w(ws2, i, 6, row["avg_unit_price"],   9, bg=bg, nf="£#,##0.00")
    w(ws2, i, 7, row["XYZ"],             10, xyz_col, bold=True, bg=bg)
    w(ws2, i, 8, row["ABC_XYZ"],         10, xyz_col, bold=True, bg=bg)


# ════════════════════════════════════════════════════════════════════
# SHEET 3 — XYZ CLASSIFICATION
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("XYZ Classification")
ws3.sheet_view.showGridLines = False
banner(ws3, "Branch 2: XYZ Classification — Demand Variability",
       "CV = σ_weekly / μ_weekly  ·  X: CV≤0.5 (predictable)  ·  Y: CV≤1.0  ·  Z: CV>1.0 (intermittent)")

section(ws3, 4, "XYZ SUMMARY TABLE")
header_row(ws3, 5,
    ["Class", "CV Threshold", "SKU Count", "% of SKUs", "Mean CV", "Median CV", "Replenishment Approach"],
    [8, 16, 12, 12, 12, 12, 44])

xyz_colors = {"X": GREEN, "Y": AMBER, "Z": CORAL}
thresholds = {"X": "≤ 0.5",  "Y": "0.5 – 1.0", "Z": "> 1.0"}
approaches = {
    "X": "Statistical ROP with tight safety stock. MRP/ERP driven. High forecast accuracy.",
    "Y": "ROP with moderate buffer. Regular forecast review. Supplier collaboration helpful.",
    "Z": "Safety stock alone insufficient. Demand sensing. VMI or consignment stock considered.",
}
for i, (_, row) in enumerate(xyz_sum.iterrows(), 6):
    ws3.row_dimensions[i].height = 28
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    col = xyz_colors.get(row["XYZ"], DGREY)
    med_cv = sku_master[sku_master["XYZ"] == row["XYZ"]]["cv"].median()
    w(ws3, i, 1, row["XYZ"],       12, col, bold=True, bg=bg)
    w(ws3, i, 2, thresholds.get(row["XYZ"], ""), 10, bg=bg)
    w(ws3, i, 3, row["sku_count"],  10, bg=bg, nf="#,##0")
    w(ws3, i, 4, row["pct_skus"],   10, col, bold=True, bg=bg, nf="0.0%")
    w(ws3, i, 5, row["cv_mean"],    10, bg=bg, nf="0.000")
    w(ws3, i, 6, float(med_cv) if not np.isnan(med_cv) else None, 10, bg=bg, nf="0.000")
    w(ws3, i, 7, approaches.get(row["XYZ"], ""), 9, DGREY, bg=bg, align="left", wrap=True)

spacer(ws3, 9)
note_row(ws3, 10,
    "⚠  KEY INSIGHT: 46% of SKUs are Z-class (CV > 1.0). This is typical for seasonal giftware / wholesale "
    "where demand is driven by events, promotions, and customer-specific bulk orders. "
    "Statistical forecasting (moving average, exponential smoothing) will systematically overstock "
    "or understock Z-class items. The correct approach is demand sensing — closer collaboration with "
    "customers, smaller order quantities, and faster replenishment cycles.",
    bg=RED_BG, color=CORAL)


# ════════════════════════════════════════════════════════════════════
# SHEET 4 — ABC-XYZ MATRIX
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("ABC-XYZ Matrix")
ws4.sheet_view.showGridLines = False
banner(ws4, "Branch 3: ABC-XYZ Combined Strategy Matrix",
       "9-cell matrix · each cell has a differentiated inventory management strategy")

section(ws4, 4, "MATRIX: SKU COUNT AND STRATEGIC ACTION PER CELL")
header_row(ws4, 5,
    ["ABC↓ / XYZ→", "X (Predictable\nCV≤0.5)", "Y (Moderate\n0.5<CV≤1.0)", "Z (Intermittent\nCV>1.0)", "ABC Row Total", "Revenue (£)", "% Revenue"],
    [18, 20, 22, 22, 16, 16, 12])

abc_row_map = {"A": CORAL, "B": AMBER, "C": SKY}
abc_xyz_counts = abc_xyz_mat.pivot_table(
    index="ABC", columns="XYZ", values="sku_count", fill_value=0
).reindex(index=["A","B","C"], columns=["X","Y","Z"], fill_value=0)

for i, abc in enumerate(["A","B","C"], 6):
    ws4.row_dimensions[i].height = 26
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    col = abc_row_map.get(abc, DGREY)
    row_total = abc_xyz_counts.loc[abc].sum()
    rev_total = sku_master[sku_master["ABC"] == abc]["total_revenue"].sum()
    w(ws4, i, 1, f"Class {abc}", 11, col, bold=True, bg=bg)
    for j, xyz in enumerate(["X","Y","Z"], 2):
        count = abc_xyz_counts.loc[abc, xyz]
        cell_label = f"{abc}{xyz}"
        cell_bg = RED_BG if cell_label in ("AZ","BZ","CZ") else GREEN_BG if cell_label in ("AX","BX","CX") else OFFWHITE
        w(ws4, i, j, int(count), 10, col, bold=True, bg=cell_bg, nf="#,##0")
    w(ws4, i, 5, int(row_total), 10, col, bold=True, bg=bg, nf="#,##0")
    w(ws4, i, 6, rev_total,      10, col, bold=True, bg=bg, nf="£#,##0")
    w(ws4, i, 7, rev_total/total_rev, 10, col, bold=True, bg=bg, nf="0.0%")

spacer(ws4, 9)
section(ws4, 10, "STRATEGIC ACTION BY CELL — MANAGEMENT INSTRUCTION SET")
header_row(ws4, 11,
    ["Cell", "SKU Count", "Revenue (£)", "Service Level", "Review Cycle", "Strategic Action"],
    [8, 12, 14, 16, 14, 58])

ACTIONS = {
    "AX": ("≥98%", "Weekly",    "Premium service. Tight ROP. MRP/ERP driven. Zero stockout tolerance."),
    "AY": ("≥95%", "Weekly",    "High service. Safety stock buffer. Weekly demand review. Preferred supplier."),
    "AZ": ("≥95%", "Daily/M'ly","MANAGEMENT ATTENTION. Demand sensing. Safety stock large. Supplier VMI/consignment."),
    "BX": ("≥95%", "Monthly",   "Standard ROP. MRP driven. Monthly review. Moderate safety stock."),
    "BY": ("≥90%", "Monthly",   "Standard ROP. Moderate buffer. Monthly review."),
    "BZ": ("≥90%", "Monthly",   "Monitor demand patterns. Collaborate with key customers on ordering intent."),
    "CX": ("85%",  "Quarterly", "Lean inventory. Min order qty. Low safety stock. Quarterly review."),
    "CY": ("85%",  "Quarterly", "Lean inventory. Consider make-to-order. Quarterly review."),
    "CZ": ("—",    "Ad hoc",    "RATIONALISATION CANDIDATE. Review with commercial team. Loyalty check required."),
}
cell_bg_map = {
    "AX": GREEN_BG, "AY": GREEN_BG, "AZ": RED_BG,
    "BX": OFFWHITE, "BY": OFFWHITE, "BZ": AMBER_BG,
    "CX": OFFWHITE, "CY": OFFWHITE, "CZ": RED_BG,
}
for i, (cell, (sl, cycle, action)) in enumerate(ACTIONS.items(), 12):
    ws4.row_dimensions[i].height = 26
    bg = cell_bg_map.get(cell, OFFWHITE)
    abc_col = {"A": CORAL, "B": AMBER, "C": SKY}.get(cell[0], DGREY)
    count_row = abc_xyz_mat[(abc_xyz_mat["ABC"]==cell[0]) & (abc_xyz_mat["XYZ"]==cell[1])]
    n = int(count_row["sku_count"].values[0]) if len(count_row) > 0 else 0
    rev = float(count_row["total_revenue"].values[0]) if len(count_row) > 0 else 0
    w(ws4, i, 1, cell,   11, abc_col, bold=True, bg=bg)
    w(ws4, i, 2, n,      10, bg=bg, nf="#,##0")
    w(ws4, i, 3, rev,    10, abc_col, bg=bg, nf="£#,##0")
    w(ws4, i, 4, sl,     10, bg=bg)
    w(ws4, i, 5, cycle,  10, bg=bg)
    w(ws4, i, 6, action,  9, CORAL if "ATTENTION" in action or "RATIONAL" in action else DGREY,
      bold=("ATTENTION" in action or "RATIONAL" in action),
      bg=bg, align="left", wrap=True)


# ════════════════════════════════════════════════════════════════════
# SHEET 5 — REORDER POINTS
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Reorder Points")
ws5.sheet_view.showGridLines = False
banner(ws5, "Branch 4: Reorder Point Model — Top 80 Revenue SKUs",
       "ROP = (μ_weekly × LT) + (z × σ_weekly × √LT)  ·  COGS fixed  ·  Service level by class")

note_row(ws5, 4,
    "FORMULA: ROP = (Average Weekly Demand × Lead Time in weeks) "
    "+ (Z-score × Weekly Demand Std Dev × √Lead Time). "
    "Lead times: A-class = 2wk, B-class = 3wk, C-class = 4wk. "
    "Z-scores: A-class = 2.054 (98% SL), B-class = 1.645 (95%), C-class = 1.281 (90%). "
    "Scope: top 80 SKUs by revenue; excludes SKUs with <3 weeks of demand data (insufficient history).",
    bg=AMBER_BG)

spacer(ws5, 5)
if not rop_df.empty:
    section(ws5, 6, "REORDER POINT TABLE — TOP 80 SKUS")
    rop_display_cols = [
        "StockCode", "Description", "ABC", "XYZ", "ABC_XYZ",
        "Annual Revenue (£)", "Avg Unit Price (£)", "Annual Units Sold",
        "avg_weekly_demand", "demand_std_weekly", "cv",
        "lead_time_weeks", "service_level_pct",
        "avg_demand_lt", "safety_stock", "reorder_point",
        "stockout_risk_flag",
    ]
    rop_headers = [
        "Stock Code", "Description", "ABC", "XYZ", "Cell",
        "Annual\nRevenue (£)", "Avg Unit\nPrice (£)", "Annual\nUnits",
        "Avg Weekly\nDemand", "Weekly\nStd Dev", "CV",
        "Lead Time\n(weeks)", "Service\nLevel %",
        "Avg Demand\nDuring LT", "Safety\nStock", "Reorder\nPoint",
        "Stockout\nRisk Flag",
    ]
    rop_widths = [12, 32, 8, 8, 8, 14, 14, 12, 14, 12, 10, 12, 12, 16, 12, 14, 14]
    header_row(ws5, 7, rop_headers, rop_widths)

    for i, (_, row) in enumerate(rop_df.iterrows(), 8):
        ws5.row_dimensions[i].height = 22
        bg = LTGREY if i % 2 == 0 else OFFWHITE
        abc_col = {"A": CORAL, "B": AMBER, "C": SKY}.get(str(row.get("ABC","C")), DGREY)
        xyz_col = {"X": GREEN, "Y": AMBER, "Z": CORAL}.get(str(row.get("XYZ","Z")), DGREY)
        is_risk  = bool(row.get("stockout_risk_flag", False))
        flag_bg  = RED_BG if is_risk else bg

        w(ws5, i, 1,  str(row.get("StockCode","")),         9, NAVY, bold=True, bg=bg)
        desc = str(row.get("Description", row.get("description","")))[:40]
        w(ws5, i, 2,  desc,                                  8, bg=bg, align="left")
        w(ws5, i, 3,  str(row.get("ABC","")),               10, abc_col, bold=True, bg=bg)
        w(ws5, i, 4,  str(row.get("XYZ","")),               10, xyz_col, bold=True, bg=bg)
        w(ws5, i, 5,  str(row.get("ABC_XYZ","")),           10, abc_col, bold=True, bg=bg)
        w(ws5, i, 6,  float(row.get("Annual Revenue (£)", 0)),   10, abc_col, bold=True, bg=bg, nf="£#,##0.00")
        w(ws5, i, 7,  float(row.get("Avg Unit Price (£)", 0)),    9, bg=bg, nf="£#,##0.00")
        w(ws5, i, 8,  int(row.get("Annual Units Sold", 0)),       9, bg=bg, nf="#,##0")
        w(ws5, i, 9,  float(row.get("avg_weekly_demand", 0)),    10, bg=bg, nf="#,##0.0")
        w(ws5, i, 10, float(row.get("demand_std_weekly", 0)),     9, bg=bg, nf="#,##0.0")
        cv_val = row.get("cv", None)
        w(ws5, i, 11, float(cv_val) if cv_val and not np.isnan(float(cv_val)) else None,
          9, CORAL if is_risk else bg, bg=bg, nf="0.000")
        w(ws5, i, 12, float(row.get("lead_time_weeks", 2)),      9, bg=bg, nf="0.0")
        w(ws5, i, 13, int(row.get("service_level_pct", 95)),     9, bg=bg, nf="0")
        w(ws5, i, 14, float(row.get("avg_demand_lt", 0)),       10, bg=bg, nf="#,##0")
        w(ws5, i, 15, float(row.get("safety_stock", 0)),        10, AMBER, bold=True, bg=bg, nf="#,##0")
        w(ws5, i, 16, float(row.get("reorder_point", 0)),       11, GREEN, bold=True, bg=bg, nf="#,##0")
        w(ws5, i, 17, "⚠ YES" if is_risk else "OK",            10,
          CORAL if is_risk else GREEN, bold=is_risk, bg=flag_bg)


# ════════════════════════════════════════════════════════════════════
# SHEET 6 — RATIONALISATION
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Rationalisation")
ws6.sheet_view.showGridLines = False
banner(ws6, "Branch 5: CZ SKU Rationalisation Analysis",
       "811 CZ SKUs identified · 3-way action split: Protect / Single Event / Rationalise")

note_row(ws6, 4,
    "ANALYST NOTE: ABC-XYZ identifies CZ SKUs as rationalisation candidates. "
    "However, the data shows most CZ SKUs serve repeat customers or represent single bulk events. "
    "ONLY 19 SKUs (£886 revenue) are true rationalisation candidates. "
    "Even these should go through commercial sign-off before discontinuation. "
    "The 136 single-event items need manual review: is the ordering customer likely to return?",
    bg=AMBER_BG)

spacer(ws6, 5)
section(ws6, 6, "CZ ACTION SUMMARY")
header_row(ws6, 7,
    ["Recommended Action", "SKU Count", "Total Revenue (£)", "Avg Revenue/SKU (£)", "Avg CV", "Avg Customers"],
    [44, 12, 18, 18, 10, 14])

if not ration_df.empty:
    action_sum = (
        ration_df.groupby("action")
        .agg(
            sku_count    = ("StockCode",   "count"),
            total_rev    = ("total_revenue","sum"),
            avg_rev      = ("total_revenue","mean"),
            avg_cv       = ("cv",           "mean"),
            avg_cust     = ("n_customers",  "mean"),
        )
        .reset_index()
    )
    action_colors = {
        "PROTECT": AMBER,
        "RATIONALISE": CORAL,
        "SINGLE": SKY,
    }
    for i, (_, row) in enumerate(action_sum.iterrows(), 8):
        ws6.row_dimensions[i].height = 30
        bg = LTGREY if i % 2 == 0 else OFFWHITE
        action_key = "PROTECT" if "PROTECT" in str(row["action"]).upper() else \
                     "RATIONALISE" if "RATIONALISE" in str(row["action"]).upper() else "SINGLE"
        col = action_colors.get(action_key, DGREY)
        w(ws6, i, 1, str(row["action"])[:55], 10, col, bold=True, bg=bg, align="left", wrap=True)
        w(ws6, i, 2, int(row["sku_count"]),   10, col, bold=True, bg=bg, nf="#,##0")
        w(ws6, i, 3, float(row["total_rev"]),  10, bg=bg, nf="£#,##0.00")
        w(ws6, i, 4, float(row["avg_rev"]),     9, bg=bg, nf="£#,##0.00")
        cv = row["avg_cv"]
        w(ws6, i, 5, float(cv) if not np.isnan(float(cv)) else None, 9, bg=bg, nf="0.000")
        w(ws6, i, 6, float(row["avg_cust"]),    9, bg=bg, nf="0.0")

    spacer(ws6, 11)
    section(ws6, 12, "CZ RATIONALISATION CANDIDATES (19 SKUs — TRUE RATIONALISE)")
    true_rational = ration_df[
        ration_df["action"].str.startswith("RATIONALISE", na=False)
    ].sort_values("total_revenue", ascending=False)
    header_row(ws6, 13,
        ["Stock Code", "Description", "Annual Revenue (£)", "No. Invoices",
         "No. Customers", "Avg Weekly Demand", "CV", "Action"],
        [12, 36, 18, 14, 14, 18, 10, 38])
    for i, (_, row) in enumerate(true_rational.iterrows(), 14):
        ws6.row_dimensions[i].height = 24
        bg = RED_BG if i % 2 == 0 else OFFWHITE
        w(ws6, i, 1, str(row.get("StockCode","")),              9, CORAL, bold=True, bg=bg)
        w(ws6, i, 2, str(row.get("description",""))[:40],       9, bg=bg, align="left")
        w(ws6, i, 3, float(row.get("total_revenue",0)),         9, CORAL, bg=bg, nf="£#,##0.00")
        w(ws6, i, 4, int(row.get("n_invoices",0)),              9, bg=bg, nf="#,##0")
        w(ws6, i, 5, int(row.get("n_customers",0)),             9, bg=bg, nf="#,##0")
        w(ws6, i, 6, float(row.get("weekly_demand_mean",0)),    9, bg=bg, nf="0.0")
        cv = row.get("cv", None)
        w(ws6, i, 7, float(cv) if cv and not np.isnan(float(cv)) else None, 9, CORAL, bold=True, bg=bg, nf="0.000")
        w(ws6, i, 8, str(row.get("action",""))[:40],            8, CORAL, bg=bg, align="left", italic=True)


# ════════════════════════════════════════════════════════════════════
# SHEET 7 — METHODOLOGY & LIMITATIONS
# ════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Methodology")
ws7.sheet_view.showGridLines = False
banner(ws7, "Methodology, Data Sources & Honest Limitations",
       "Transparency on approach, assumptions, and what the model cannot tell you")

section(ws7, 4, "DATA SOURCE")
data_rows = [
    ("Dataset name",   "UCI Online Retail Dataset"),
    ("Hosted at",      "archive.ics.uci.edu/dataset/352/online+retail"),
    ("Nature of data", "Real UK-based, non-store online retailer (giftware/wholesale). All-occasion gifts."),
    ("Raw records",    "541,909 invoice line items across 12 months (Dec 2010 – Dec 2011)"),
    ("Clean records",  "525,462 after removing cancellations, negative quantities, invalid stock codes"),
    ("Unique SKUs",    "3,788 after cleaning"),
    ("Citation",       "Daqing Chen, Sai Laing Sain, Kun Guo (2012). Data mining for the online retail "
                       "industry: A case study of RFM model-based customer segmentation. Journal of "
                       "Database Marketing and Customer Strategy Management, 19(3)."),
]
header_row(ws7, 5, ["Item", "Detail"], [22, 72])
for i, (item, detail) in enumerate(data_rows, 6):
    ws7.row_dimensions[i].height = 26
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws7, i, 1, item,   10, NAVY, bold=True, bg=bg, align="left")
    ws7.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
    c = ws7.cell(row=i, column=2, value=detail)
    c.font = font(10); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ba(c)

spacer(ws7, 13)
section(ws7, 14, "METHODOLOGY — ABC-XYZ FRAMEWORK")
method_rows = [
    ("ABC thresholds",    "A = 0-80% cumulative revenue · B = 80-95% · C = 95-100%. "
                          "Standard thresholds. Alternatives (70/90 or 75/95) exist but 80/95 is widely adopted."),
    ("XYZ metric",        "Coefficient of Variation (CV) = σ_weekly / μ_weekly. Weekly demand used to reduce "
                          "day-of-week noise and align with typical replenishment cycles."),
    ("XYZ thresholds",    "X: CV ≤ 0.5 · Y: CV ≤ 1.0 · Z: CV > 1.0. Literature ranges vary (0.5/1.0 and "
                          "0.3/0.7 are both used). This project uses 0.5/1.0 which is common in UK retail contexts."),
    ("ROP formula",       "ROP = (μ_weekly × LT) + (z × σ_weekly × √LT). Standard stochastic ROP under normally "
                          "distributed demand. The normal distribution assumption is simplifying — actual demand "
                          "may be right-skewed, particularly for Z-class items."),
    ("Lead time",         "Assumed by class (A:2wk, B:3wk, C:4wk). These are reasonable proxies. Real "
                          "implementation requires actual supplier lead time data from the ERP system."),
    ("Service levels",    "A-class 98%, B-class 95%, C-class 90%. Differentiated by commercial importance. "
                          "These targets should be validated against customer SLA agreements."),
]
header_row(ws7, 15, ["Element", "Detail"], [22, 72])
for i, (elem, detail) in enumerate(method_rows, 16):
    ws7.row_dimensions[i].height = 32
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws7, i, 1, elem, 10, NAVY, bold=True, bg=bg, align="left")
    ws7.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
    c = ws7.cell(row=i, column=2, value=detail)
    c.font = font(10); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ba(c)

spacer(ws7, 22)
section(ws7, 23, "HONEST LIMITATIONS", bg=CORAL)
limitations = [
    ("No stock-on-hand data", "The ROP model computes when to reorder, but we have no current stock levels. "
                              "The ROP table gives the trigger point — the operations team needs to compare "
                              "against live ERP stock positions."),
    ("Demand normality",      "Safety stock formula assumes normally distributed demand. Z-class items "
                              "frequently violate this. Safety stock for Z-class SKUs should be reviewed "
                              "manually and may need to be set using empirical percentiles rather than σ."),
    ("Lead time variability", "Lead times are assumed fixed. In practice, lead time variability (σ_LT) "
                              "can significantly increase required safety stock. If supplier reliability "
                              "is low, adjust: SS = z × √(LT × σ_d² + μ_d² × σ_LT²)."),
    ("Single-period data",    "Analysis covers one year (Dec 2010 – Dec 2011). Seasonal patterns are "
                              "embedded but not explicitly decomposed. Giftware demand has strong Christmas "
                              "seasonality — annual averages will understate Q4 requirements."),
    ("CZ rationalisation",   "The rationalisation candidates (19 SKUs) should be reviewed with the commercial "
                              "team before any action. 'Numbers say remove it' is not sufficient justification."),
]
header_row(ws7, 24, ["Limitation", "What It Means for the Model"], [24, 70], bg=CORAL)
for i, (lim, detail) in enumerate(limitations, 25):
    ws7.row_dimensions[i].height = 34
    bg = RED_BG if i % 2 == 0 else OFFWHITE
    w(ws7, i, 1, lim, 10, CORAL, bold=True, bg=bg, align="left")
    ws7.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
    c = ws7.cell(row=i, column=2, value=detail)
    c.font = font(10); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ba(c)

# Footer
ws7.merge_cells("A31:I31")
cf = ws7.cell(row=31, column=1,
    value="Vaishnavi Bhor · MSc Business Analytics, University of Manchester · "
          "vbhor207@gmail.com · linkedin.com/in/vaishnavi-bhor-business-analyst · vbho.github.io/portfolio")
cf.font = font(10, SKY); cf.fill = fill(OFFWHITE)
cf.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ba(cf); ws7.row_dimensions[31].height = 20

# ── Save ──────────────────────────────────────────────────────────────────────
outpath = OUT / "Inventory_Optimisation_VaishnaviBhor.xlsx"
wb.save(outpath)
print(f"\n✓  Workbook saved  →  {outpath}")
print(f"   Sheets: {', '.join(ws.title for ws in wb.worksheets)}")
